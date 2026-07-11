#!/usr/bin/env python3
"""
Script para extrair peças e looks do Excel e gerar JSON para a app.

USE ESTE SCRIPT QUANDO:
- Adicionar novas peças ao Excel
- Atualizar dados das peças existentes
- Criar novos looks

COMO USAR:
1. Coloque este arquivo na mesma pasta que seu Excel
2. Abra terminal/cmd na pasta
3. Digite: python extrair_dados.py
4. Ele gera um novo dados_guarda_roupa.json
5. Coloque esse JSON na pasta da app
6. Recarregue a página da app
"""

import openpyxl
import json
from datetime import datetime
import sys
import os
import re

def valor_texto(valor):
    if valor is None:
        return ''
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return str(valor).strip()

def normalizar_id(valor):
    texto = valor_texto(valor).upper()
    return texto if re.fullmatch(r'[A-Z]{1,4}\d{4}', texto) and texto != 'ID0000' else ''

def nome_campo(header, col_idx):
    return valor_texto(header) or f'col_{col_idx + 1}'

def campo_valido(valor):
    texto = valor_texto(valor)
    return bool(texto and texto.lower() != 'na')

def extrair_campos_linha(row, headers, indices):
    campos = []
    for col_idx in indices:
        if col_idx >= len(row):
            continue
        valor = valor_texto(row[col_idx])
        if not campo_valido(valor):
            continue
        campos.append({
            'campo': nome_campo(headers[col_idx] if col_idx < len(headers) else '', col_idx),
            'valor': valor,
            'coluna': col_idx + 1,
        })
    return campos

def extrair_ids_linha(row, headers, indices):
    itens = []
    for col_idx in indices:
        if col_idx >= len(row):
            continue
        peca_id = normalizar_id(row[col_idx])
        if not peca_id:
            continue
        itens.append({
            'grupo': nome_campo(headers[col_idx] if col_idx < len(headers) else '', col_idx),
            'id': peca_id,
            'foto': f'fotos/{peca_id}.webp',
        })
    return itens

def extrair_mapa_combinacoes_nao_permitidas(wb):
    """
    Lê a aba Categorias e retorna descrições para códigos CD0001...
    Na planilha, esse bloco fica nas colunas AY:BA.
    """
    try:
        ws = wb['Categorias']
    except KeyError:
        return {}

    mapa = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=51, max_col=53, values_only=True):
        codigo = valor_texto(row[0]).upper()
        if not codigo:
            continue

        mapa[codigo] = {
            'codigo': codigo,
            'descricao': valor_texto(row[1]) or codigo,
            'grupo': valor_texto(row[2]),
            'foto': f'fotos/categorias/{codigo}.webp',
        }

    return mapa

def buscar_combinacao_nao_permitida(mapa, valor):
    texto = valor_texto(valor).upper()
    if not texto:
        return {}

    if texto in mapa:
        return mapa[texto]

    for info in mapa.values():
        if valor_texto(info.get('descricao')).upper() == texto:
            return info

    return {}

def extrair_mapa_ocasiões(wb):
    """
    Lê a aba Categorias e retorna descrições para os códigos de ocasião.
    Na planilha, esse bloco fica nas colunas E:I.
    """
    try:
        ws = wb['Categorias']
    except KeyError:
        return {}

    mapa = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=9, values_only=True):
        codigo = valor_texto(row[0])
        if not codigo:
            continue

        mapa[codigo] = {
            'codigo': codigo,
            'descricao': valor_texto(row[1]) or codigo,
            'local': valor_texto(row[2]),
            'tipo': valor_texto(row[3]),
            'data_revisao': valor_texto(row[4]),
        }

    return mapa

def extrair_mapa_climas(wb):
    """
    Lê a aba Categorias e retorna descrições para os códigos de clima.
    Na planilha, esse bloco fica nas colunas A:C.
    """
    try:
        ws = wb['Categorias']
    except KeyError:
        return {}

    mapa = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
        codigo = valor_texto(row[0])
        if not codigo:
            continue

        mapa[codigo] = {
            'codigo': codigo,
            'descricao': valor_texto(row[1]) or codigo,
            'temperatura': valor_texto(row[2]),
        }

    return mapa

def extrair_linhas_dimensao(ws, min_col, max_col, nomes=None):
    """Extrai um bloco vertical da aba Categorias, ignorando linhas vazias."""
    headers = [valor_texto(cell.value) for cell in ws[1][min_col - 1:max_col]]
    chaves = nomes or [
        re.sub(r'[^a-z0-9_]+', '_', header.lower()).strip('_') or f'coluna_{idx + min_col}'
        for idx, header in enumerate(headers)
    ]
    linhas = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min_col, max_col=max_col, values_only=True):
        valores = [valor_texto(valor) for valor in row]
        if not any(valores):
            continue
        linhas.append({chaves[idx]: valor for idx, valor in enumerate(valores) if idx < len(chaves)})
    return linhas


def extrair_dimensoes(wb, mapa_climas, mapa_ocasioes):
    """Carrega todas as dimensoes mantidas na aba Categorias."""
    try:
        ws = wb['Categorias']
    except KeyError:
        return {}

    # J:Q guarda a necessidade por clima; R guarda o total por ocasiao.
    codigos_clima_meta = [valor_texto(ws.cell(1, coluna).value) for coluna in range(10, 18)]
    for linha in range(2, ws.max_row + 1):
        codigo = valor_texto(ws.cell(linha, 5).value)
        if not codigo or codigo not in mapa_ocasioes:
            continue
        metas = {}
        for coluna, clima in zip(range(10, 18), codigos_clima_meta):
            valor = ws.cell(linha, coluna).value
            if clima and valor not in (None, ''):
                metas[clima] = valor
        mapa_ocasioes[codigo]['quantidades_necessarias'] = metas
        total_planilha = ws.cell(linha, 18).value
        mapa_ocasioes[codigo]['total_necessario'] = total_planilha if isinstance(total_planilha, (int, float)) else sum(
            valor for valor in metas.values() if isinstance(valor, (int, float))
        )

    tipos_peca = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=32, max_col=49, values_only=True):
        codigo, tipo, grupo = [valor_texto(valor) for valor in row[:3]]
        if not tipo:
            continue
        tipos_peca.append({
            'codigo': codigo,
            'tipo': tipo,
            'grupo': grupo,
            'combinacoes': [valor_texto(valor) for valor in row[3:] if valor_texto(valor)],
        })

    locais = [
        {'valor': valor_texto(ws.cell(linha, 30).value)}
        for linha in range(2, ws.max_row + 1)
        if valor_texto(ws.cell(linha, 30).value)
    ]
    aquecimentos = [
        {'valor': valor_texto(ws.cell(linha, 80).value)}
        for linha in range(2, ws.max_row + 1)
        if valor_texto(ws.cell(linha, 80).value)
    ]

    return {
        'climas': list(mapa_climas.values()),
        'ocasioes': list(mapa_ocasioes.values()),
        'situacoes_look': extrair_linhas_dimensao(ws, 20, 21, ['codigo', 'valor']),
        'categorias_look': extrair_linhas_dimensao(ws, 23, 25, ['codigo', 'categoria', 'indicador']),
        'utilizacoes_look': extrair_linhas_dimensao(ws, 27, 28, ['codigo', 'valor']),
        'locais': locais,
        'tipos_peca': tipos_peca,
        'tipos_combinacao': extrair_linhas_dimensao(ws, 51, 53, ['codigo', 'tipo', 'grupo']),
        'conservacoes_peca': extrair_linhas_dimensao(ws, 55, 56, ['codigo', 'valor']),
        'reposicoes_peca': extrair_linhas_dimensao(ws, 58, 59, ['codigo', 'valor']),
        'funcoes_peca': extrair_linhas_dimensao(ws, 61, 62, ['codigo', 'valor']),
        'padronagens_peca': extrair_linhas_dimensao(ws, 64, 65, ['codigo', 'valor']),
        'cores_detalhe': extrair_linhas_dimensao(ws, 67, 69, ['codigo', 'cor_detalhe', 'cor']),
        'cores_peca': extrair_linhas_dimensao(ws, 71, 72, ['codigo', 'valor']),
        'cores_tons': extrair_linhas_dimensao(ws, 74, 75, ['cor', 'tom']),
        'tons_peca': extrair_linhas_dimensao(ws, 77, 78, ['codigo', 'valor']),
        'aquecimentos_peca': aquecimentos,
        'formalidades_peca': extrair_linhas_dimensao(ws, 82, 83, ['codigo', 'valor']),
        'tendencias_peca': extrair_linhas_dimensao(ws, 85, 86, ['codigo', 'valor']),
        'alocacoes_peca': extrair_linhas_dimensao(ws, 88, 89, ['codigo', 'valor']),
        'situacoes_peca': extrair_linhas_dimensao(ws, 91, 92, ['codigo', 'valor']),
        'utilizacoes_peca': extrair_linhas_dimensao(ws, 94, 95, ['codigo', 'valor']),
        'tipos_subtipos_peca': extrair_linhas_dimensao(ws, 97, 98, ['tipo', 'subtipo']),
    }


def normalizar_categoria(valor):
    import unicodedata
    texto = unicodedata.normalize('NFKD', valor_texto(valor).casefold())
    return ''.join(char for char in texto if not unicodedata.combining(char)).strip()


def validar_dimensoes(pecas, looks, dimensoes):
    """Compara os bancos com as listas da aba Categorias."""
    faltantes = {}

    def validar(nome, valores, permitidos):
        permitidos_norm = {normalizar_categoria(valor) for valor in permitidos if campo_valido(valor)}
        contagens = {}
        originais = {}
        for valor in valores:
            if not campo_valido(valor):
                continue
            chave = normalizar_categoria(valor)
            if chave and chave not in permitidos_norm:
                contagens[chave] = contagens.get(chave, 0) + 1
                originais.setdefault(chave, valor_texto(valor))
        if contagens:
            faltantes[nome] = [
                {'valor': originais[chave], 'ocorrencias': contagens[chave]}
                for chave in sorted(contagens, key=lambda item: (-contagens[item], item))
            ]

    def valores_dim(nome, campo='valor'):
        return [item.get(campo, '') for item in dimensoes.get(nome, [])]

    mapa_campos_peca = {
        'tipo_peça': ('tipo', valores_dim('tipos_peca', 'tipo')),
        'função_peça': ('funcao', valores_dim('funcoes_peca')),
        'subtipo_peça': ('subtipo', valores_dim('tipos_subtipos_peca', 'subtipo')),
        'aquecimento_peça': ('nivel_aquecimento', valores_dim('aquecimentos_peca')),
        'padronagem_peça': ('padronagem', valores_dim('padronagens_peca')),
        'cor_detalhe_peça': ('cor_detalhe', valores_dim('cores_detalhe', 'cor_detalhe')),
        'tom_peça': ('tom', valores_dim('tons_peca')),
        'utilização_peça': ('utilizacao', valores_dim('utilizacoes_peca')),
        'formalidade_peça': ('formalidade', valores_dim('formalidades_peca')),
        'tendência_peça': ('tendencia', valores_dim('tendencias_peca')),
        'local_peça': ('local', valores_dim('locais')),
        'alocação_peça': ('alocacao', valores_dim('alocacoes_peca')),
        'situação_peça': ('situacao', valores_dim('situacoes_peca')),
        'conservação_peça': ('conservacao', valores_dim('conservacoes_peca')),
        'reposição_peça': ('reposicao', valores_dim('reposicoes_peca')),
    }
    for nome, (campo, permitidos) in mapa_campos_peca.items():
        validar(nome, (peca.get(campo) for peca in pecas.values()), permitidos)

    pares_permitidos = {
        (normalizar_categoria(item.get('tipo')), normalizar_categoria(item.get('subtipo')))
        for item in dimensoes.get('tipos_subtipos_peca', [])
    }
    pares_faltantes = {}
    for peca in pecas.values():
        if not campo_valido(peca.get('subtipo')):
            continue
        par = (normalizar_categoria(peca.get('tipo')), normalizar_categoria(peca.get('subtipo')))
        if par not in pares_permitidos:
            rotulo = f"{valor_texto(peca.get('tipo'))} → {valor_texto(peca.get('subtipo'))}"
            pares_faltantes[rotulo] = pares_faltantes.get(rotulo, 0) + 1
    if pares_faltantes:
        faltantes['relação_tipo_subtipo_peça'] = [
            {'valor': valor, 'ocorrencias': quantidade}
            for valor, quantidade in sorted(pares_faltantes.items(), key=lambda item: (-item[1], item[0]))
        ]

    cor_por_detalhe = {
        normalizar_categoria(item.get('cor_detalhe')): item.get('cor', '')
        for item in dimensoes.get('cores_detalhe', [])
    }
    pares_cor_tom = {
        (normalizar_categoria(item.get('cor')), normalizar_categoria(item.get('tom')))
        for item in dimensoes.get('cores_tons', [])
    }
    relacoes_cor_tom = {}
    for peca in pecas.values():
        cor = cor_por_detalhe.get(normalizar_categoria(peca.get('cor_detalhe')), '')
        tom = peca.get('tom')
        if not campo_valido(cor) or not campo_valido(tom):
            continue
        if (normalizar_categoria(cor), normalizar_categoria(tom)) not in pares_cor_tom:
            rotulo = f"{cor} → {valor_texto(tom)}"
            relacoes_cor_tom[rotulo] = relacoes_cor_tom.get(rotulo, 0) + 1
    if relacoes_cor_tom:
        faltantes['relação_cor_tom_peça'] = [
            {'valor': valor, 'ocorrencias': quantidade}
            for valor, quantidade in sorted(relacoes_cor_tom.items(), key=lambda item: (-item[1], item[0]))
        ]

    validar('situação_look', (look.get('situacao') for look in looks.values()), valores_dim('situacoes_look'))
    validar('indicador_look', (look.get('indicador') for look in looks.values()), valores_dim('categorias_look', 'indicador'))
    validar('utilização_look', (look.get('utilizacao') for look in looks.values()), valores_dim('utilizacoes_look'))
    validar('local_look', (look.get('local') for look in looks.values()), valores_dim('locais'))
    validar('clima_look', (look.get('clima') for look in looks.values()), [item.get('codigo') for item in dimensoes.get('climas', [])])
    validar(
        'ocasião_look',
        (ocasiao.get('codigo') for look in looks.values() for ocasiao in look.get('ocasioes', [])),
        [item.get('codigo') for item in dimensoes.get('ocasioes', [])],
    )

    return {
        'valido': not faltantes,
        'campos_com_opcoes_faltantes': len(faltantes),
        'faltantes': faltantes,
    }


def salvar_relatorio_validacao(validacao, nome_arquivo='RELATORIO_CATEGORIAS_FALTANTES.md'):
    linhas = ['# Validação da aba Categorias', '']
    faltantes = validacao.get('faltantes', {})
    if not faltantes:
        linhas.append('Todos os valores dos bancos de looks e peças constam na aba Categorias.')
    else:
        linhas.append('Adicione à aba Categorias as opções abaixo para cobrir os valores já usados nos bancos.')
        for campo, itens in faltantes.items():
            linhas.extend(['', f'## {campo}', ''])
            linhas.extend(f"- {item['valor']} ({item['ocorrencias']} ocorrência(s))" for item in itens)
    with open(nome_arquivo, 'w', encoding='utf-8') as arquivo:
        arquivo.write('\n'.join(linhas) + '\n')


def aquecimento_peca(pecas, peca_id):
    valor = pecas.get(peca_id, {}).get('nivel_aquecimento')
    texto = valor_texto(valor)
    return texto if texto and texto != 'na' else None

def local_peca(pecas, peca_id):
    valor = pecas.get(peca_id, {}).get('local')
    texto = valor_texto(valor)
    return texto if texto and texto != 'na' else None

def utilizacao_peca(pecas, peca_id):
    valor = pecas.get(peca_id, {}).get('utilizacao')
    texto = valor_texto(valor)
    return texto if texto and texto != 'na' else None

def em(valor, opcoes):
    return valor in opcoes

def calcular_local_look(situacao, loc1, loc2, loc3):
    situacao = valor_texto(situacao)
    locs = [valor_texto(loc) or None for loc in [loc1, loc2, loc3]]

    if 'virtual' in locs or situacao == 'excluído':
        return 'virtual'

    if locs[0] and (locs[0] == locs[1] or locs[1] is None) and (locs[0] == locs[2] or locs[2] is None):
        return locs[0]

    return 'misto'

def calcular_utilizacao_look(indicador, util1, util2, util3):
    indicador = valor_texto(indicador)
    u1, u2, u3 = [valor_texto(util) or None for util in [util1, util2, util3]]

    if indicador in ['LL', 'PL']:
        return 'under'

    if u1 == 'casa' and em(u2, ['casa', None]) and em(u3, ['casa', None]):
        return 'casa'

    if u1 == 'sair' and em(u2, ['sair', None]) and em(u3, ['sair', None]):
        return 'sair'

    if u1 == 'sair' and ((u2 is not None and u2 != 'sair') or (u3 is not None and u3 != 'sair')):
        return 'mix'

    if u1 == 'casa' and ((u2 is not None and u2 != 'casa') or (u3 is not None and u3 != 'casa')):
        return 'mix'

    return None

def calcular_clima_look(indicador, aquece1, aquece2, aquece3):
    indicador = valor_texto(indicador)
    a1 = valor_texto(aquece1) or None
    a2 = valor_texto(aquece2) or None
    a3 = valor_texto(aquece3) or None

    if a1 == "1" and em(a2, ["1", None]) and a3 is None:
        return "1"

    if a1 == "2" and em(a2, ["1", None]) and a3 is None:
        return "2"
    if indicador != "TL" and em(a1, ["1", "2"]) and a2 == "2" and a3 is None:
        return "2"
    if indicador != "TL" and em(a1, ["1", "2"]) and em(a2, ["1", "2", None]) and a3 == "2":
        return "2"
    if indicador == "TL" and a1 == "1" and a2 == "2" and a3 is None:
        return "2"

    if indicador != "TL" and em(a1, ["1", "2"]) and em(a2, ["1", "2", None]) and a3 == "3":
        return "3"
    if indicador == "TL" and a1 == "1" and em(a2, ["1", "2"]) and a3 == "3":
        return "3"
    if indicador == "TL" and a1 == "2" and a2 == "2" and a3 is None:
        return "3"

    if a1 == "3" and em(a2, ["1", "2"]) and em(a3, ["3", None]):
        return "4"
    if indicador != "TL" and em(a1, ["1", "2"]) and em(a2, ["1", "2", None]) and a3 == "4":
        return "4"
    if indicador == "TL" and a1 == "2" and em(a2, ["1", "2"]) and a3 == "3":
        return "4"

    if a1 == "4" and em(a2, ["2", "1"]) and em(a3, [None, "2", "3"]):
        return "5"
    if indicador != "TL" and em(a1, ["1", "2"]) and em(a2, ["1", "2", None]) and a3 == "5":
        return "5"
    if indicador != "TL" and a1 == "3" and em(a2, ["1", "2", None]) and a3 == "4":
        return "5"
    if a1 == "3" and a2 == "5" and a3 is None:
        return "5"
    if indicador == "TL" and em(a1, ["1", "2"]) and a2 == "2" and a3 == "4":
        return "5"

    if indicador != "TL" and a1 == "5" and em(a2, ["1", "2"]) and em(a3, [None, "2", "3"]):
        return "6"
    if indicador != "TL" and a1 == "3" and a2 == "2" and a3 == "5":
        return "6"
    if indicador != "TL" and em(a1, ["1", "2"]) and a2 == "2" and a3 == "6":
        return "6"
    if indicador != "TL" and em(a1, ["4", "5"]) and a2 == "2" and a3 == "4":
        return "6"
    if indicador != "TL" and a1 == "4" and a2 == "5" and em(a3, [None, "2"]):
        return "6"
    if indicador == "TL" and em(a1, ["1", "2", "3"]) and a2 == "2" and a3 == "5":
        return "6"
    if indicador == "TL" and a1 == "5" and a2 == "2" and a3 is None:
        return "6"
    if indicador == "TL" and a1 == "2" and a2 == "5" and em(a3, ["3", None]):
        return "6"

    if a1 == "6" and a2 == "2" and em(a3, [None, "2"]):
        return "7"
    if indicador != "TL" and a1 == "3" and a2 == "2" and a3 == "6":
        return "7"
    if indicador != "TL" and em(a1, ["4", "5"]) and em(a2, ["2", "5"]) and a3 == "5":
        return "7"
    if indicador != "TL" and a1 == "5" and a2 == "5" and em(a3, [None, "2"]):
        return "7"
    if indicador != "TL" and a1 == "3" and a2 == "5" and a3 == "5":
        return "7"
    if indicador == "TL" and a1 == "2" and a2 == "5" and a3 == "4":
        return "7"
    if indicador == "TL" and em(a1, ["3", "4"]) and a2 == "5" and a3 is None:
        return "7"

    if indicador != "TL" and em(a1, ["4", "5"]) and em(a2, ["2", "5"]) and a3 == "6":
        return "8"
    if indicador != "TL" and a1 == "6" and a2 == "5" and em(a3, [None, "2"]):
        return "8"
    if indicador != "TL" and a1 == "3" and a2 == "5" and a3 == "6":
        return "8"
    if indicador == "TL" and a1 == "2" and a2 == "5" and em(a3, ["5", "6"]):
        return "8"
    if indicador == "TL" and em(a1, ["5", "6"]) and a2 == "5" and a3 is None:
        return "8"

    if indicador == "PRL" and a1 == "0" and em(a2, ["0", "1", None]) and em(a3, ["1", "2", None]):
        return "1"

    return None

def extrair_dados(arquivo_excel):
    """
    Extrai dados do Excel e gera JSON para a app.
    
    Argumentos:
        arquivo_excel (str): Caminho do arquivo .xlsx ou .xlsm
    
    Retorna:
        dict: Dicionário com pecas, looks e ocasioes
    """
    
    print(f"📖 Abrindo arquivo: {arquivo_excel}")
    print("-" * 50)
    
    try:
        # Abrir o Excel
        wb = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        print(f"❌ Erro: Arquivo '{arquivo_excel}' não encontrado!")
        print(f"Coloque o arquivo .xlsx/.xlsm na pasta do script.")
        return None
    except Exception as e:
        print(f"❌ Erro ao abrir arquivo: {e}")
        return None

    # ==================== EXTRAIR PEÇAS ====================
    print("\n🔄 Extraindo peças...")
    
    try:
        ws_pecas = wb['BD peças']
    except KeyError:
        print("❌ Erro: Aba 'BD peças' não encontrada!")
        return None
    
    pecas = {}
    headers_pecas = [valor_texto(cell.value) for cell in ws_pecas[1]]
    campos_extra_idx = [11, 12, 13, 15, 16, 17, 18, 20, 21, 22, 23, 24, 25, 26, 27, 28]
    acessorios_idx = range(37, 61) # AL:BI
    combinacoes_nao_permitidas_idx = range(61, 69) # BJ:BQ
    mapa_combinacoes_nao_permitidas = extrair_mapa_combinacoes_nao_permitidas(wb)
    
    # Estrutura esperada (confira com seu Excel):
    # Coluna A: ID
    # Coluna D: Tipo
    # Coluna E: Função
    # Coluna F: Subtipo
    # coluna G: nivel_aquecimento
    # Coluna H: Padronagem
    # Coluna I: Cor detalhe
    # Coluna J: Tom
    # Coluna K: Utilização
    # Coluna O: Local
    # Coluna Q: Situação
    
    for row_idx, row in enumerate(ws_pecas.iter_rows(min_row=2, max_row=ws_pecas.max_row, values_only=True), start=2):
        try:
            id_peca = row[0]  # Coluna A
            
            # Pular linhas vazias ou cabeçalho repetido
            if not id_peca or id_peca == 'ID' or not isinstance(id_peca, str):
                continue
            
            tipo = str(row[3]) if row[3] else 'na'  # Coluna D
            funcao = str(row[4]) if row[4] else 'na'      # Coluna E
            subtipo = str(row[5]) if row[5] else 'na'  # Coluna F
            nivel_aquecimento = str(row[6]) if row[6] else 'na'  # Coluna G
            padronagem = str(row[7]) if row[7] else 'na'  # Coluna H
            cor_detalhe = str(row[8]) if row[8] else 'na'  # Coluna I
            tom = str(row[9]) if row[9] else 'na'  # Coluna J
            utilizacao = str(row[10]) if row[10] else 'na'  # Coluna K
            formalidade = str(row[12]) if row[12] else 'na'  # Coluna M
            tendencia = str(row[13]) if row[13] else 'na'  # Coluna N
            local = str(row[14]) if row[14] else 'na'  # Coluna O
            alocacao = str(row[15]) if row[15] else 'na'  # Coluna P
            situacao = str(row[16]) if row[16] else 'na'  # Coluna Q
            conservacao = str(row[17]) if row[17] else 'na'  # Coluna R
            reposicao = str(row[18]) if row[18] else 'na'  # Coluna S
            info_fotos = valor_texto(row[25]) if len(row) > 25 else ''  # Coluna Z
            combinacoes = valor_texto(row[26]) if len(row) > 26 else ''  # Coluna AA
            detalhes = extrair_campos_linha(row, headers_pecas, campos_extra_idx)
            acessorios = extrair_ids_linha(row, headers_pecas, acessorios_idx)
            combinacoes_nao_permitidas = []

            for col_idx in combinacoes_nao_permitidas_idx:
                if col_idx >= len(row):
                    continue

                valor_restricao = valor_texto(row[col_idx])
                if not campo_valido(valor_restricao):
                    continue

                info = buscar_combinacao_nao_permitida(mapa_combinacoes_nao_permitidas, valor_restricao)
                codigo = info.get('codigo', valor_restricao.upper())
                combinacoes_nao_permitidas.append({
                    'grupo': nome_campo(headers_pecas[col_idx] if col_idx < len(headers_pecas) else '', col_idx),
                    'codigo': codigo,
                    'descricao': info.get('descricao', codigo),
                    'categoria': info.get('grupo', ''),
                    'foto': info.get('foto', f'fotos/categorias/{codigo}.webp'),
                })

            pecas[id_peca] = {
                'id': id_peca,
                'tipo': tipo.strip(),
                'funcao': funcao.strip(),
                'subtipo': subtipo.strip(),
                'nivel_aquecimento': nivel_aquecimento.strip(),
                'padronagem': padronagem.strip(),
                'cor_detalhe': cor_detalhe.strip(),
                'tom': tom.strip(),
                'utilizacao': utilizacao.strip(),
                'formalidade': formalidade.strip(),
                'tendencia': tendencia.strip(),
                'local': local.strip(),
                'alocacao': alocacao.strip(),
                'situacao': situacao.strip(),
                'conservacao': conservacao.strip(),
                'reposicao': reposicao.strip(),
                'info_fotos': info_fotos,
                'combinacoes': combinacoes,
                'detalhes': detalhes,
                'acessorios': acessorios,
                'combinacoes_nao_permitidas': combinacoes_nao_permitidas
            }
        
        except (IndexError, TypeError) as e:
            # Pular erros de linha mal formatada
            continue
    
    print(f"✅ {len(pecas)} peças extraídas com sucesso!")
    
    # ==================== EXTRAIR LOOKS ====================
    print("\n🔄 Extraindo looks...")
    
    try:
        ws_looks = wb['BD looks']
    except KeyError:
        print("⚠️  Aviso: Aba 'BD looks' não encontrada. Continuando sem looks...")
        ws_looks = None
    
    looks = {}
    mapa_ocasiões = extrair_mapa_ocasiões(wb)
    mapa_climas = extrair_mapa_climas(wb)
    dimensoes = extrair_dimensoes(wb, mapa_climas, mapa_ocasiões)
    cores_por_detalhe = {
        normalizar_categoria(item.get('cor_detalhe')): item.get('cor', '')
        for item in dimensoes.get('cores_detalhe', [])
    }
    for peca in pecas.values():
        peca['cor'] = cores_por_detalhe.get(normalizar_categoria(peca.get('cor_detalhe')), '')
    
    if ws_looks:
        headers = [valor_texto(cell.value) for cell in ws_looks[1]]
        campos_basicos_idx = range(0, 16)   # A:P
        pecas_sugeridas_idx = range(16, 35) # Q:AI
        ocasioes_idx = range(35, 76)        # AJ:BX
        
        for row_idx, row in enumerate(ws_looks.iter_rows(min_row=2, max_row=ws_looks.max_row, values_only=True), start=2):
            try:
                id_look = normalizar_id(row[0])  # Coluna A
                
                if not id_look:
                    continue
                
                situacao = valor_texto(row[6]) or 'ativo'  # Coluna G
                
                pecas_look = []
                for col_idx in [1, 2, 3]:
                    peca_id = normalizar_id(row[col_idx] if col_idx < len(row) else None)
                    if peca_id:
                        pecas_look.append(peca_id)

                basicos = {}
                for col_idx in campos_basicos_idx:
                    if col_idx >= len(row):
                        continue
                    header = headers[col_idx] or f'col_{col_idx + 1}'
                    basicos[header] = valor_texto(row[col_idx])

                pecas_sugeridas = []
                for col_idx in pecas_sugeridas_idx:
                    if col_idx >= len(row):
                        continue
                    peca_id = normalizar_id(row[col_idx])
                    if peca_id:
                        pecas_sugeridas.append({
                            'grupo': headers[col_idx] or f'col_{col_idx + 1}',
                            'id': peca_id,
                        })

                ocasioes = []
                for col_idx in ocasioes_idx:
                    if col_idx >= len(row) or not row[col_idx]:
                        continue
                    codigo = headers[col_idx]
                    if not codigo:
                        continue
                    info = mapa_ocasiões.get(codigo, {})
                    ocasioes.append({
                        'codigo': codigo,
                        'descricao': info.get('descricao', codigo),
                        'local': info.get('local', ''),
                        'tipo': info.get('tipo', ''),
                    })

                clima = valor_texto(row[76]) if len(row) > 76 else ''
                utilizacao = valor_texto(row[77]) if len(row) > 77 else ''
                local = valor_texto(row[78]) if len(row) > 78 else ''
                aquecimentos = [
                    aquecimento_peca(pecas, pecas_look[0]) if len(pecas_look) > 0 else None,
                    aquecimento_peca(pecas, pecas_look[1]) if len(pecas_look) > 1 else None,
                    aquecimento_peca(pecas, pecas_look[2]) if len(pecas_look) > 2 else None,
                ]
                locais_pecas = [
                    local_peca(pecas, pecas_look[0]) if len(pecas_look) > 0 else None,
                    local_peca(pecas, pecas_look[1]) if len(pecas_look) > 1 else None,
                    local_peca(pecas, pecas_look[2]) if len(pecas_look) > 2 else None,
                ]
                utilizacoes_pecas = [
                    utilizacao_peca(pecas, pecas_look[0]) if len(pecas_look) > 0 else None,
                    utilizacao_peca(pecas, pecas_look[1]) if len(pecas_look) > 1 else None,
                    utilizacao_peca(pecas, pecas_look[2]) if len(pecas_look) > 2 else None,
                ]
                clima_calc = calcular_clima_look(valor_texto(row[10]), *aquecimentos)
                clima_final = clima_calc or clima
                clima_info = mapa_climas.get(clima_final, {'codigo': clima_final, 'descricao': clima_final, 'temperatura': ''}) if clima_final else {}
                local_calc = calcular_local_look(situacao, *locais_pecas)
                utilizacao_calc = calcular_utilizacao_look(valor_texto(row[10]), *utilizacoes_pecas)
                
                if pecas_look:
                    looks[id_look] = {
                        'id': id_look,
                        'pecas': pecas_look,
                        'pecas_sugeridas': pecas_sugeridas,
                        'ocasioes': ocasioes,
                        'ocasiao': ', '.join([o['descricao'] for o in ocasioes]) or 'não especificada',
                        'situacao': situacao.strip(),
                        'indicador': valor_texto(row[10]) if len(row) > 10 else '',
                        'categoria': next((
                            item.get('categoria', '') for item in dimensoes.get('categorias_look', [])
                            if normalizar_categoria(item.get('indicador')) == normalizar_categoria(valor_texto(row[10]))
                        ), ''),
                        'clima': clima_final,
                        'clima_planilha': clima,
                        'clima_calc': clima_calc or '',
                        'clima_info': clima_info,
                        'aquecimentos': aquecimentos,
                        'local_calc': local_calc,
                        'locais_pecas': locais_pecas,
                        'utilizacao_calc': utilizacao_calc or '',
                        'utilizacoes_pecas': utilizacoes_pecas,
                        'utilizacao': utilizacao,
                        'local': local,
                        'basicos': basicos,
                        'foto': f"fotos/{id_look.rstrip('0123456789')}/{id_look}.webp",
                    }
            
            except (IndexError, TypeError) as e:
                continue
        
        print(f"✅ {len(looks)} looks extraídos com sucesso!")
    
    # ==================== RETORNAR DADOS ====================
    
    validacao_dimensoes = validar_dimensoes(pecas, looks, dimensoes)

    dados = {
        'pecas': pecas,
        'looks': looks,
        'ocasioes': mapa_ocasiões,
        'climas': mapa_climas,
        'combinacoes_nao_permitidas': mapa_combinacoes_nao_permitidas,
        'dimensoes': dimensoes,
        'validacao_dimensoes': validacao_dimensoes,
        'ultima_atualizacao': datetime.now().isoformat(),
        'total_pecas': len(pecas),
        'total_looks': len(looks),
    }
    
    return dados


def salvar_json(dados, nome_arquivo='dados_guarda_roupa.json'):
    """
    Salva os dados em um arquivo JSON.
    
    Argumentos:
        dados (dict): Dados a salvar
        nome_arquivo (str): Nome do arquivo de saída
    """
    
    print(f"\n💾 Salvando em {nome_arquivo}...")
    
    try:
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        
        # Calcular tamanho do arquivo
        tamanho_kb = os.path.getsize(nome_arquivo) / 1024
        
        print(f"✅ Arquivo '{nome_arquivo}' criado com sucesso!")
        print(f"   Tamanho: {tamanho_kb:.1f} KB")
    
    except Exception as e:
        print(f"❌ Erro ao salvar arquivo: {e}")


def mostrar_resumo(dados):
    """
    Mostra um resumo dos dados extraídos.
    
    Argumentos:
        dados (dict): Dados extraídos
    """
    
    print("\n" + "=" * 50)
    print("📊 RESUMO DOS DADOS EXTRAÍDOS")
    print("=" * 50)
    
    print(f"\n📍 Peças:")
    print(f"   Total: {dados['total_pecas']}")
    
    if dados['pecas']:
        # Mostrar alguns exemplos
        exemplos = list(dados['pecas'].items())[:3]
        for id_peca, peca in exemplos:
            print(f"   • {peca['id']}: {peca['tipo']} ({peca['cor_detalhe']})")
        
        if len(dados['pecas']) > 3:
            print(f"   ... e mais {len(dados['pecas']) - 3}")
    
    print(f"\n💄 Looks:")
    print(f"   Total: {dados['total_looks']}")
    
    if dados['looks']:
        # Mostrar alguns exemplos
        exemplos = list(dados['looks'].items())[:3]
        for id_look, look in exemplos:
            pecas_nomes = ', '.join([
                dados['pecas'][pid]['tipo'] if pid in dados['pecas'] else pid
                for pid in look['pecas']
            ])
            print(f"   • {look['id']}: {pecas_nomes}")
        
        if len(dados['looks']) > 3:
            print(f"   ... e mais {len(dados['looks']) - 3}")
    
    print(f"\n📅 Última atualização: {dados['ultima_atualizacao']}")
    print("\n" + "=" * 50)


def main():
    """
    Função principal - ponto de entrada do script.
    """
    
    print("\n" + "=" * 50)
    print("👗 EXTRATOR DE DADOS - GUARDA-ROUPA")
    print("=" * 50)
    
    # Caminho fixo do arquivo Excel
    arquivo_excel = r"C:\Users\livia\OneDrive\0 Organização e Planejamento\0001 Controles\0305_Vestuário.xlsm"
    
    print(f"\n📂 Usando arquivo: {arquivo_excel}\n")
    
    # Extrair dados
    dados = extrair_dados(arquivo_excel)
    
    if not dados:
        print("\n❌ Falha ao extrair dados!")
        return
    
    # Mostrar resumo
    mostrar_resumo(dados)
    
    # Salvar JSON
    salvar_json(dados)
    salvar_relatorio_validacao(dados.get('validacao_dimensoes', {}))
    
    print("\n✅ Pronto! Agora:")
    print("   1. Copie 'dados_guarda_roupa.json' para a pasta da app")
    print("   2. Recarregue a página da app (Ctrl+F5)")
    print("   3. Seus dados estão atualizados!")
    
    print("\n" + "=" * 50 + "\n")


if __name__ == '__main__':
    main()
